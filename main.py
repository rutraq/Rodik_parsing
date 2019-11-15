from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import requests
from lxml import html
import os
import shutil
import re
from transliterate import translit
import openpyxl


class Driver:
    def __init__(self):
        options = webdriver.FirefoxOptions()
        options.add_argument('headless')
        load = DesiredCapabilities().FIREFOX
        load["pageLoadStrategy"] = "eager"
        self.driver = webdriver.Firefox(desired_capabilities=load,
                                        executable_path=r"C:/Users/Lenovo/Desktop/geckodriver.exe",
                                        firefox_options=options)


class Parsing(Driver):
    def __init__(self):
        super().__init__()
        self.url = "http://www.tools.by/?q=kat/920359/923254"
        self.driver.get(self.url)
        self.wait = WebDriverWait(self.driver, 3)
        self.find_products(self.url)

    def find_products(self, url):
        # count = 4
        # photo = 3
        r = requests.get(url)
        tree = html.fromstring(r.content)
        products = tree.xpath('//*/tbody/tr/td[3]/a[1]/@href')
        del products[0]
        if os.path.exists("photos"):
            shutil.rmtree("photos")
        os.mkdir("photos")
        i = 1

        length = len(products)
        count_percents = length / 100
        percents = 100
        clear_line = "--------------------------------------------------"
        progress_line = ""
        progress_plus = 2
        count = 1

        for product in products:
            # self.information_excel(product, count, info)
            # count += 1
            # photo += 1
            r = requests.get(product)
            tree = html.fromstring(r.content)
            src = tree.xpath('//*[@id="show-img"]/@src')
            try:
                photo = requests.get(src[0])
            except IndexError:
                print(src)
                break
            with open("photos/{0}.jpg".format(i), 'wb') as out:
                out.write(photo.content)

            more_photos = tree.xpath('//*[@id="small-img-roll"]/div[2]/img/@src')
            if len(more_photos) != 0:
                self.additional_photos(i, more_photos, tree)

            if (i / count_percents) >= progress_plus:
                while progress_plus <= (i / count_percents):
                    progress_plus += 2
                    progress_line += "#"
                    count += 1
                print("\r{0}% |{1}{2}| {3} of {4}".format(round(i / count_percents, 1), progress_line,
                                                          clear_line[count:-1], i, length),
                      end="")
            else:
                print("\r{0}% |{1}{2}|  {3} of {4}".format(round(i / count_percents, 1), progress_line,
                                                           clear_line[count:-1], i, length),
                      end="")
            i += 1

    @staticmethod
    def additional_photos(i, more_photos, tree):
        j = 1
        path = "http://www.tools.by/newkatfiles/products/"
        try:
            while True:
                photos = re.search(r'\d+_', more_photos[0])[0]
                src = "{0}{1}{2}.jpg".format(path, photos, j)
                photo = requests.get(src)
                with open("photos/{0}_{1}.jpg".format(i, j), 'wb') as out:
                    out.write(photo.content)
                j += 1
                more_photos = tree.xpath('//*[@id="small-img-roll"]/div[{0}]/img/@src'.format(j + 1))
                if len(more_photos) == 0:
                    break
        except IndexError:
            print(IndexError)
        except TypeError:
            print(TypeError)

    def information_excel(self, url, count, photo):
        self.driver.get(url)
        full_info = dict()
        name = self.wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#product h1")))
        full_info['name'] = name.text[7:-8]
        print(full_info['name'])
        try:
            price = self.wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "price_block_n")))
            full_info['price'] = price.text[12:-5]
        except TimeoutException:
            full_info['price'] = 0

        try:
            description = self.wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "product_text")))
            full_info['description'] = description.text
            full_info['description'] = ''
            desc = description.text.split('\n')
            for tabulation in desc:
                full_info['description'] += '<p>' + tabulation + '</p>'
        except TimeoutException:
            full_info['description'] = ''

        page = requests.get(url)
        x = re.findall(r'<b>Страна изготовления:\W+\w+.\s+[А-Я][а-я]+', page.text)
        manufacturer = x[0]
        manufacturer = manufacturer[28::]

        wb = openpyxl.load_workbook(filename='1.xlsx')
        sheet = wb['Products']
        productsID = sheet.cell(row=count - 1, column=1).value
        productsID += 1
        sheet.cell(row=count, column=1).value = productsID
        sheet.cell(row=count, column=2).value = full_info['name']
        sheet.cell(row=count, column=3).value = '2005'
        sheet.cell(row=count, column=4).value = sheet.cell(row=count - 1, column=4).value
        sheet.cell(row=count, column=5).value = sheet.cell(row=count - 1, column=5).value
        sheet.cell(row=count, column=6).value = sheet.cell(row=count - 1, column=6).value
        sheet.cell(row=count, column=7).value = sheet.cell(row=count - 1, column=7).value
        sheet.cell(row=count, column=8).value = sheet.cell(row=count - 1, column=8).value
        sheet.cell(row=count, column=9).value = sheet.cell(row=count - 1, column=9).value
        sheet.cell(row=count, column=10).value = sheet.cell(row=count - 1, column=10).value
        if full_info['price'] == 0:
            sheet.cell(row=count, column=11).value = 0
        else:
            sheet.cell(row=count, column=11).value = 99
        sheet.cell(row=count, column=12).value = str(productsID - 1) + '-01'
        sheet.cell(row=count, column=13).value = sheet.cell(row=count - 1, column=13).value
        sheet.cell(row=count, column=14).value = 'catalog/bathhouse/{0}.jpg'.format(photo)
        sheet.cell(row=count, column=15).value = sheet.cell(row=count - 1, column=15).value
        sheet.cell(row=count, column=16).value = full_info['price']

        sheet.cell(row=count, column=17).value = sheet.cell(row=count - 1, column=17).value
        sheet.cell(row=count, column=18).value = sheet.cell(row=count - 1, column=18).value
        sheet.cell(row=count, column=19).value = sheet.cell(row=count - 1, column=19).value
        sheet.cell(row=count, column=20).value = sheet.cell(row=count - 1, column=20).value
        sheet.cell(row=count, column=21).value = sheet.cell(row=count - 1, column=21).value
        sheet.cell(row=count, column=22).value = sheet.cell(row=count - 1, column=22).value
        sheet.cell(row=count, column=23).value = sheet.cell(row=count - 1, column=23).value
        sheet.cell(row=count, column=24).value = sheet.cell(row=count - 1, column=24).value
        sheet.cell(row=count, column=25).value = sheet.cell(row=count - 1, column=25).value
        sheet.cell(row=count, column=26).value = sheet.cell(row=count - 1, column=26).value
        sheet.cell(row=count, column=27).value = sheet.cell(row=count - 1, column=27).value
        sheet.cell(row=count, column=28).value = sheet.cell(row=count - 1, column=28).value

        sheet.cell(row=count, column=29).value = translit(
            full_info['name'].replace('"', '').replace(',', '-').replace(' ', '-').replace('--', '-'), reversed=True)

        sheet.cell(row=count, column=30).value = full_info['description']
        sheet.cell(row=count, column=31).value = full_info['name']
        sheet.cell(row=count, column=32).value = full_info['name']

        sheet.cell(row=count, column=33).value = sheet.cell(row=count - 1, column=33).value
        sheet.cell(row=count, column=34).value = sheet.cell(row=count - 1, column=34).value
        sheet.cell(row=count, column=35).value = sheet.cell(row=count - 1, column=35).value
        sheet.cell(row=count, column=36).value = sheet.cell(row=count - 1, column=36).value
        sheet.cell(row=count, column=37).value = sheet.cell(row=count - 1, column=37).value
        sheet.cell(row=count, column=38).value = sheet.cell(row=count - 1, column=38).value
        sheet.cell(row=count, column=39).value = sheet.cell(row=count - 1, column=39).value
        sheet.cell(row=count, column=40).value = sheet.cell(row=count - 1, column=40).value
        sheet.cell(row=count, column=41).value = sheet.cell(row=count - 1, column=41).value
        wb.save('1.xlsx')
        sheet = wb['ProductAttributes']
        sheet.cell(row=count, column=1).value = productsID
        sheet.cell(row=count, column=2).value = 'Характеристики товара'
        sheet.cell(row=count, column=3).value = 'Страна производства'
        sheet.cell(row=count, column=4).value = manufacturer
        wb.save('1.xlsx')
        print('Сохранено')

    @staticmethod
    def check_photos():
        import os

        files = os.listdir("photos")

        photo_number = 1
        photos = []  # можешь удалить после доработки

        for i in range(len(files)):
            src = "{0}.jpg".format(photo_number)

            if src in files:
                photos.append(src)  # загрузка главного фото
            else:
                break

            j = 1
            while True:
                additional_src = "{0}_{1}.jpg".format(photo_number, j)
                if additional_src in files:
                    photos.append(additional_src)  # загрузка дополнительного фото, вставь код для additional_images
                    j += 1
                else:
                    break

            photo_number += 1


if __name__ == "__main__":
    info = Parsing()
