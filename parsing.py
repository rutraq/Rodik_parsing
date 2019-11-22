from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import requests
from lxml import html
import os
import re
from transliterate import translit
import openpyxl


class Driver:
    def __init__(self):
        options = webdriver.FirefoxOptions()
        options.add_argument('headless')
        load = DesiredCapabilities().FIREFOX
        load["pageLoadStrategy"] = "eager"
        self.driver = webdriver.Firefox(desired_capabilities=load,
                                        executable_path=r"C:/Users/evgen/Desktop/geckodriver.exe",
                                        firefox_options=options)


class Parsing(Driver):
    def __init__(self):
        super().__init__()
        self.url = "http://www.tools.by/?q=kat/3341/1008277"
        self.len_id_product = list()
        self.text_id_product = list()
        self.count_id = 0
        self.driver.get(self.url)
        self.wait = WebDriverWait(self.driver, 3)
        self.find_products(self.url)

    def find_products(self, url):
        print('Начало')
        count_process = 1
        id_product = self.wait.until(EC.visibility_of_all_elements_located((By.CLASS_NAME, "div_artikul")))
        for id in id_product:
            self.len_id_product.append(len(str(id.text)) + 2)
            self.text_id_product.append(str(id.text))
        count = 1052
        photo = 1
        r = requests.get(url)
        tree = html.fromstring(r.content)
        products = tree.xpath('//*/tbody/tr/td[3]/a[1]/@href')
        del products[0]
        for product in products:
            print('Объект {0}'.format(count_process))
            self.information_excel(product, count, photo)
            count += 1
            photo += 1
            count_process += 1

    def information_excel(self, url, count, photo):
        self.driver.get(url)
        full_info = dict()
        try:
            name = self.wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#product h1")))
            name_id_delete = self.text_id_product[self.count_id]
            full_info['name'] = name.text.replace(name_id_delete, "").replace("()", " ")
            print(full_info['name'])
        except:
            full_info['name'] = ' '
            name = self.wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#product h1")))
            name_id_delete = "(" + self.text_id_product[self.count_id] + ")"
            print(name.text.replace(name_id_delete, ""))
        try:
            price = self.wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "price_block_n")))
            full_info['price'] = price.text[12:-5]
        except TimeoutException:
            full_info['price'] = 0

        try:
            description = self.wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "product_text")))
            full_info['description'] = description.text
            full_info['description'] = ''
            desc = description.text.split('\n')
            for tabulation in desc:
                full_info['description'] += '<p>' + tabulation + '</p>'
        except TimeoutException:
            full_info['description'] = ''

        page = requests.get(url)
        x = re.findall(r'<b>Страна изготовления:\W+\w+.\s+[А-Я][а-я]+', page.text)
        try:
            manufacturer = x[0]
            manufacturer = manufacturer[28::]
        except:
            manufacturer = ''

        wb = openpyxl.load_workbook(filename='1.xlsx')
        sheet = wb['Products']

        productsID = int(sheet.cell(row=count - 1, column=1).value)
        productsID += 1
        sheet.cell(row=count, column=1).value = productsID
        sheet.cell(row=count, column=2).value = full_info['name']
        sheet.cell(row=count, column=3).value = '2030'
        sheet.cell(row=count, column=4).value = sheet.cell(row=count - 1, column=4).value
        sheet.cell(row=count, column=5).value = sheet.cell(row=count - 1, column=5).value
        sheet.cell(row=count, column=6).value = sheet.cell(row=count - 1, column=6).value
        sheet.cell(row=count, column=7).value = sheet.cell(row=count - 1, column=7).value
        sheet.cell(row=count, column=8).value = sheet.cell(row=count - 1, column=8).value
        sheet.cell(row=count, column=9).value = sheet.cell(row=count - 1, column=9).value
        sheet.cell(row=count, column=10).value = sheet.cell(row=count - 1, column=10).value
        if full_info['price'] == 0:
            sheet.cell(row=count, column=11).value = 0
        else:
            sheet.cell(row=count, column=11).value = 99
        sheet.cell(row=count, column=12).value = str(productsID + 1) + '-01'
        sheet.cell(row=count, column=13).value = sheet.cell(row=count - 1, column=13).value
        sheet.cell(row=count, column=14).value = 'catalog/For_household_goods/Scissors/{0}.jpg'.format(photo)
        sheet.cell(row=count, column=15).value = sheet.cell(row=count - 1, column=15).value
        sheet.cell(row=count, column=16).value = full_info['price']
        sheet.cell(row=count, column=17).value = sheet.cell(row=count - 1, column=17).value
        sheet.cell(row=count, column=18).value = sheet.cell(row=count - 1, column=18).value
        sheet.cell(row=count, column=19).value = sheet.cell(row=count - 1, column=19).value
        sheet.cell(row=count, column=20).value = sheet.cell(row=count - 1, column=20).value
        sheet.cell(row=count, column=21).value = sheet.cell(row=count - 1, column=21).value
        sheet.cell(row=count, column=22).value = sheet.cell(row=count - 1, column=22).value
        sheet.cell(row=count, column=23).value = sheet.cell(row=count - 1, column=23).value
        sheet.cell(row=count, column=24).value = sheet.cell(row=count - 1, column=24).value
        sheet.cell(row=count, column=25).value = sheet.cell(row=count - 1, column=25).value
        sheet.cell(row=count, column=26).value = sheet.cell(row=count - 1, column=26).value
        sheet.cell(row=count, column=27).value = sheet.cell(row=count - 1, column=27).value
        sheet.cell(row=count, column=28).value = sheet.cell(row=count - 1, column=28).value
        sheet.cell(row=count, column=29).value = translit(
            full_info['name'].replace('"', '').replace(',', '-').replace(' ', '-').replace('--', '-').replace('.',
                                                                                                              '').replace(
                '+', '').replace('(', '').replace(')', '').replace('/', ''), reversed=True)
        try:
            sheet.cell(row=count, column=30).value = str(full_info['description'])
        except:
            sheet.cell(row=count, column=30).value = ' '
        sheet.cell(row=count, column=31).value = full_info['name']
        sheet.cell(row=count, column=32).value = full_info['name']
        sheet.cell(row=count, column=33).value = sheet.cell(row=count - 1, column=33).value
        sheet.cell(row=count, column=34).value = sheet.cell(row=count - 1, column=34).value
        sheet.cell(row=count, column=35).value = sheet.cell(row=count - 1, column=35).value
        sheet.cell(row=count, column=36).value = sheet.cell(row=count - 1, column=36).value
        sheet.cell(row=count, column=37).value = sheet.cell(row=count - 1, column=37).value
        sheet.cell(row=count, column=38).value = sheet.cell(row=count - 1, column=38).value
        sheet.cell(row=count, column=39).value = sheet.cell(row=count - 1, column=39).value
        sheet.cell(row=count, column=40).value = sheet.cell(row=count - 1, column=40).value
        sheet.cell(row=count, column=41).value = sheet.cell(row=count - 1, column=41).value
        wb.save('1.xlsx')
        sheet = wb['ProductAttributes']
        sheet.cell(row=count, column=1).value = productsID
        sheet.cell(row=count, column=2).value = 'Характеристики товара'
        sheet.cell(row=count, column=3).value = 'Страна производства'
        sheet.cell(row=count, column=4).value = manufacturer
        wb.save('1.xlsx')
        files = os.listdir("photos")
        j = 1
        sheet = wb['AdditionalImages']
        dopphoto = int(sheet.cell(row=2, column=4).value) + 1
        while True:
            additional_src = "{0}_{1}.jpg".format(photo, j)
            if additional_src in files:
                sheet = wb['AdditionalImages']
                sheet.cell(row=dopphoto, column=1).value = productsID
                sheet.cell(row=dopphoto, column=2).value = 'catalog/For_household_goods/Scissors/{0}'.format(
                    additional_src)
                sheet.cell(row=dopphoto, column=3).value = 0
                sheet.cell(row=2, column=4).value = dopphoto
                print(additional_src)
                print(dopphoto)
                dopphoto += 1
                j += 1
                wb.save('1.xlsx')
            else:
                break
        print('Сохранено')
        self.count_id += 1


if __name__ == "__main__":
    info = Parsing()
